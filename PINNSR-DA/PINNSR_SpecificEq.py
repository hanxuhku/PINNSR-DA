# -*- coding: utf-8 -*-
"""
PINNSR Framework for case-specific governing equation discovery.

======================================================================================================================
Created on Mon Jun 24 20:32:37 2024

Inspired by:
1. Chen, Z., Liu, Y., & Sun, H. (2021). Physics-informed learning of governing equations from scarce data. 
   Nature Communications, 12(1), 6136.
2. Rudy, S. H., Brunton, S. L., Proctor, J. L., & Kutz, J. N. (2017). Data-driven discovery of partial 
   differential equations. Science Advances, 3(4), e1602614.       

Novel Contributions:
1. Enhanced PINN-SR architecture with input-output structures tailored for granular rheological conditions.
2. Adaptive collocation sampling strategy to prioritize physical information in high-gradient regions 
   induced by shear reversal.
======================================================================================================================

@author: Han Xu
"""
#%% Import required libraries
import os
import pandas as pd
import time
import math
import numpy as np
import torch
import torch.nn as nn
import matplotlib.pyplot as plt
import matplotlib as mpl
import scipy.io
from scipy.io import loadmat, savemat
from scipy.spatial import distance
from scipy.stats import gaussian_kde
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, NamedStyle

# Custom Utilities
from utils.GenerateData_AdaptionCollection import GenerateData
from utils.ReadFolder import ReadFolder

# Global Timing
GLOBAL_START_TIME = time.time()

# Set up device (GPU if available)
device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
#%%
# Number of SINDy candidate terms
num_coe = 23

# Initialize loss histories to record convergence
# Adam loss history
loss_history_Adam = np.array([0])
loss_u_history_Adam = np.array([0])
loss_f_history_Adam = np.array([0])
loss_lambda_history_Adam = np.array([0])
lambda_history_Adam = np.zeros((num_coe, 1))
loss_history_Adam_val = np.array([0])
loss_u_history_Adam_val = np.array([0])
loss_f_history_Adam_val = np.array([0])

# STRidge loss history
loss_history_STRidge = np.array([0])
loss_f_history_STRidge = np.array([0])
loss_lambda_history_STRidge = np.array([0])
optimaltol_history = np.array([0])
tol_history_STRidge = np.array([0])
lambda_normalized_history_STRidge = np.zeros((num_coe, 1))

lambda_history_STRidge = np.zeros((num_coe, 1))
ridge_append_counter_STRidge = np.array([0])

# Loss histories for pretraining
loss_history_Pretrain = np.array([0])
loss_u_history_Pretrain = np.array([0])
loss_f_history_Pretrain = np.array([0])
loss_lambda_history_Pretrain = np.array([0])
loss_history_val_Pretrain = np.array([0])
loss_u_history_val_Pretrain = np.array([0])
loss_f_history_val_Pretrain = np.array([0])
step_Pretrain = 0

lambda_history_Pretrain = np.zeros((num_coe, 1))

# Set random seeds for reproducibility
np.random.seed(1234)
torch.manual_seed(1234)
torch.cuda.manual_seed_all(42)

class PINNSR:
    """
    Physics-Informed Neural Network (PINN) with Sparse Identification of Nonlinear Dynamics (SINDy)
    
    This class implements a hybrid model combining neural networks with physical constraints
    to identify governing equations from data using the SINDy approach.
    """
    
    def __init__(self, Data_Train, Data_Collection, Data_Validation, layers, lb, ub, 
                 name_list, project_root, collection_multiple, directory_path, log_flag=False):
        """
        Initialize the Physics-Informed Neural Network
        
        Parameters:
        -----------
        Data_Train : numpy array
            Training dataset with shape (n_samples, 5) containing [t, γ, γ_t, γ_tt, σ]
        Data_Collection : numpy array
            Collocation points with shape (n_points, 4) containing [t, γ, γ_t, γ_tt]
        Data_Validation : numpy array
            Validation dataset with shape (n_samples, 5) containing [t, γ, γ_t, γ_tt, σ]
        layers : list
            List specifying the number of neurons in each layer of the neural network
        lb : numpy array
            Lower bounds of the input space [t_min, γ_min]
        ub : numpy array
            Upper bounds of the input space [t_max, γ_max]
        name_list : list
            List of filenames for data loading
        project_root : str
            Root directory of the project
        collection_multiple : int
            Multiplier for collocation point generation
        directory_path : str
            Path to data directory
        log_flag : bool
            Flag indicating whether to use logarithmic scaling
        """
        self.lb = torch.tensor(lb, dtype=torch.float32, device=device)
        self.ub = torch.tensor(ub, dtype=torch.float32, device=device)
        self.layers = layers
        self.it = 0  # Iteration counter for ADO loop
        
        # Initialize neural network weights and biases
        self.weights, self.biases = self.initialize_NN(layers)
        
        # Initialize SINDy coefficients
        self.lambda1 = nn.Parameter(torch.zeros(num_coe, 1, dtype=torch.float32, device=device, requires_grad=True))
        
        # Training data
        self.x = torch.tensor(Data_Train[:, 5:6], dtype=torch.float32, device=device, requires_grad=True)  # gamma0
        self.t = torch.tensor(Data_Train[:, 0:1], dtype=torch.float32, device=device, requires_grad=True)  # time
        self.u = torch.tensor(Data_Train[:, 4:5], dtype=torch.float32, device=device, requires_grad=True)  # sigma (σ)
        
        # Collocation points for physics loss
        self.x_f = torch.tensor(Data_Collection[:, 4:5], dtype=torch.float32, device=device, requires_grad=True)
        self.t_f = torch.tensor(Data_Collection[:, 0:1], dtype=torch.float32, device=device, requires_grad=True)
        
        # Validation data
        self.x_val = torch.tensor(Data_Validation[:, 5:6], dtype=torch.float32, device=device, requires_grad=True)
        self.t_val = torch.tensor(Data_Validation[:, 0:1], dtype=torch.float32, device=device, requires_grad=True)
        self.u_val = torch.tensor(Data_Validation[:, 4:5], dtype=torch.float32, device=device, requires_grad=True)
        
        # Strain information matrices
        self.G_f = torch.tensor(Data_Collection[:, 1:5], dtype=torch.float32, device=device, requires_grad=True)
        self.G_val = torch.tensor(Data_Validation[:, 1:5], dtype=torch.float32, device=device, requires_grad=True)
        
        # Collect all parameters
        self.parameters = self.weights + self.biases + [self.lambda1]
        
        # Optimizers
        var_list_1 = self.biases + self.weights
        self.optimizer_Adam = torch.optim.Adam(var_list_1, lr=1e-3, betas=(0.99, 0.9), eps=1e-8)
        self.optimizer_LBFGS = torch.optim.LBFGS(var_list_1, lr=1, max_iter=1000, max_eval=1000, 
                                                tolerance_grad=1e-7, tolerance_change=1e-9, 
                                                history_size=50, line_search_fn="strong_wolfe")
        self.optimizer_Pretrain = torch.optim.LBFGS(self.parameters, max_iter=20000, history_size=50, 
                                                  max_eval=20000, tolerance_grad=1e-7, 
                                                  tolerance_change=1e-9, line_search_fn="strong_wolfe")
        
        self.loss_f_coeff = 0.01  # Coefficient for physics loss term

    def initialize_NN(self, layers):
        """
        Initialize neural network weights using Xavier initialization
        
        Parameters:
        -----------
        layers : list
            List specifying the number of neurons in each layer
            
        Returns:
        --------
        weights : list
            List of weight tensors for each layer
        biases : list
            List of bias tensors for each layer
        """
        weights = []
        biases = []
        num_layers = len(layers)
        for l in range(num_layers - 1):
            W = self.xavier_init(size=[layers[l], layers[l + 1]])
            b = torch.zeros([1, layers[l + 1]], dtype=torch.float32, device=device, requires_grad=True)
            weights.append(W)
            biases.append(b)
        return weights, biases
    
    def xavier_init(self, size):
        """
        Xavier initialization for neural network weights
        
        Parameters:
        -----------
        size : tuple
            Size of the weight matrix (input_dim, output_dim)
            
        Returns:
        --------
        torch.Tensor
            Initialized weight matrix
        """
        in_dim = size[0]
        out_dim = size[1]
        xavier_stddev = np.sqrt(2 / (in_dim + out_dim))
        return torch.nn.Parameter(torch.randn([in_dim, out_dim], dtype=torch.float32, device=device) * xavier_stddev)

    def neural_net(self, x, t, weights, biases):
        """
        Forward pass through the neural network
        
        Parameters:
        -----------
        x : torch.Tensor
            Input tensor for γ (strain)
        t : torch.Tensor
            Input tensor for time
        weights : list
            List of weight tensors
        biases : list
            List of bias tensors
            
        Returns:
        --------
        torch.Tensor
            Network output (predicted σ)
        """
        X = torch.cat((x, t), dim=1)
        num_layers = len(weights) + 1
        # Normalize input to [-1, 1] range
        H = 2.0 * (X - self.lb) / (self.ub - self.lb) - 1.0
        
        # Forward pass through hidden layers with tanh activation
        for l in range(num_layers - 2):
            W = weights[l]
            b = biases[l]
            H = torch.tanh(torch.add(torch.matmul(H, W), b))
        
        # Output layer (no activation)
        W = weights[-1]
        b = biases[-1]
        Y = torch.add(torch.matmul(H, W), b)
        return Y
    
    def net_u(self, x, t):
        """
        Predict σ using the neural network
        
        Parameters:
        -----------
        x : torch.Tensor
            Input tensor for γ (strain)
        t : torch.Tensor
            Input tensor for time
            
        Returns:
        --------
        torch.Tensor
            Predicted σ value
        """
        u = self.neural_net(x, t, self.weights, self.biases)
        return u
    
    def net_f(self, x, t, G):
        """
        Compute the SINDy library and the residual of the identified equation
        
        Parameters:
        -----------
        x : torch.Tensor
            Input tensor for γ (strain)
        t : torch.Tensor
            Input tensor for time
        G : torch.Tensor
            Matrix containing strain information [γ, γ', γ'']
            
        Returns:
        --------
        f : torch.Tensor
            Residual of the equation (Ax - b)
        Phi : torch.Tensor
            SINDy library matrix
        u_t : torch.Tensor
            Time derivative of σ (∂σ/∂t)
        """
        u = self.net_u(x, t)
        # Compute time derivative using automatic differentiation
        u_t = torch.autograd.grad(u, t, grad_outputs=torch.ones_like(u), create_graph=True)[0]
        
        # Extract strain components
        g = G[:, 0:1]       # γ
        g_t = G[:, 1:2]     # γ'
        g_tt = G[:, 2:3]    # γ''
        abs_g_t = torch.abs(G[:, 1:2])   # |γ'|
        abs_g_tt = torch.abs(G[:, 2:3])  # |γ''|
        
        # Create ones tensor matching dimensions
        ones_tensor = torch.ones((x.shape[0], 1), dtype=torch.float32, device=device)
        
        # Construct SINDy library (feature matrix)
        Phi = torch.cat([ones_tensor, u, g_t, g_tt, u*u, g_t*g_t, g_tt*g_tt, abs_g_t, abs_g_tt,
                         u*g_t, u*g_tt, u*abs_g_t, u*abs_g_tt, abs_g_t*g_t, abs_g_t*g_tt,
                         (u*u)*g_t, (u*u)*g_tt, (u*u)*abs_g_t, (u*u)*abs_g_tt, g_t*(g_tt*g_tt), 
                         g_tt*(g_t*g_t), u_t*g_t, u_t*abs_g_t], 1)
        
        Phi = Phi.to(torch.float32)
        
        # Description of each term in the SINDy library
        self.library_description = ['1', 'σ', 'γ_t', 'γ_tt', 'σ^2', 'γ_t^2', 'γ_tt^2', '|γ_t|', '|γ_tt|',
                                    'σ*γ_t', 'σ*γ_tt', 'σ*|γ_t|', 'σ*|γ_tt|', '|γ_t|*γ_t', '|γ_t|*γ_tt',
                                    'σ^2*γ_t', 'σ^2*γ_tt', 'σ^2*|γ_t|', 'σ^2*|γ_tt|', 'γ_t*γ_tt^2', 
                                    'γ_t^2*γ_tt', 'σ_t*γ_t', 'σ_t*|γ_t|']
        
        # Residual of the equation: Φλ - ∂σ/∂t = 0
        f = torch.matmul(Phi, self.lambda1) - u_t
        return f, Phi, u_t

    def loss_function(self):
        """
        Compute the composite loss function including data loss, physics loss, and regularization
        
        Returns:
        --------
        loss : torch.Tensor
            Total training loss
        loss_u : torch.Tensor
            Data mismatch loss
        loss_f : torch.Tensor
            Physics loss
        loss_lambda : torch.Tensor
            Regularization loss for SINDy coefficients
        loss_val : torch.Tensor
            Total validation loss
        loss_u_val : torch.Tensor
            Validation data mismatch loss
        loss_f_val : torch.Tensor
            Validation physics loss
        """
        # Predictions on training data
        u_pred = self.net_u(self.x, self.t)
        
        # Physics loss on collocation points
        f_pred, _, _ = self.net_f(self.x_f, self.t_f, self.G_f)
        
        # Predictions on validation data
        u_val_pred = self.net_u(self.x_val, self.t_val)
        f_val_pred, _, _ = self.net_f(self.x_val, self.t_val, self.G_val)
        
        # Compute loss components
        loss_u = torch.mean((self.u - u_pred) ** 2)  # Data mismatch loss
        loss_f = self.loss_f_coeff * torch.mean(f_pred ** 2)  # Physics loss
        loss_lambda = 1e-7 * torch.norm(self.lambda1, p=1)  # L1 regularization for sparsity
        
        # Validation loss components
        loss_u_val = torch.mean((self.u_val - u_val_pred) ** 2)
        loss_f_val = self.loss_f_coeff * torch.mean(f_val_pred ** 2)
        
        # Total loss (logarithmic scaling for better optimization)
        loss = torch.log(loss_u + loss_f + loss_lambda)
        loss_val = torch.log(loss_u_val + loss_f_val)
        
        return loss, loss_u, loss_f, loss_lambda, loss_val, loss_u_val, loss_f_val

    def callback_Pretrain(self, loss, loss_u, loss_f, loss_lambda, loss_val, loss_u_val, loss_f_val, lamu):
        """
        Callback function to record loss histories during pretraining
        
        Parameters:
        -----------
        loss : torch.Tensor
            Total training loss
        loss_u : torch.Tensor
            Data mismatch loss
        loss_f : torch.Tensor
            Physics loss
        loss_lambda : torch.Tensor
            Regularization loss
        loss_val : torch.Tensor
            Total validation loss
        loss_u_val : torch.Tensor
            Validation data mismatch loss
        loss_f_val : torch.Tensor
            Validation physics loss
        lamu : torch.Tensor
            Current SINDy coefficients
        """
        global step_Pretrain
        step_Pretrain += 1
        if step_Pretrain % 10 == 0:
            print(f'Step: {step_Pretrain}, log Loss: {loss:.2e}, loss_u: {loss_u:.2e}, '
                  f'loss_f: {loss_f:.2e}, loss_lambda: {loss_lambda:.2e}')
            
            # Update global loss histories
            global loss_history_Pretrain, loss_u_history_Pretrain, loss_f_history_Pretrain
            global loss_lambda_history_Pretrain, loss_history_val_Pretrain
            global loss_u_history_val_Pretrain, loss_f_history_val_Pretrain, lambda_history_Pretrain
            
            loss_history_Pretrain = np.append(loss_history_Pretrain, loss.item())
            loss_u_history_Pretrain = np.append(loss_u_history_Pretrain, loss_u.item())
            loss_f_history_Pretrain = np.append(loss_f_history_Pretrain, loss_f.item())
            loss_lambda_history_Pretrain = np.append(loss_lambda_history_Pretrain, loss_lambda.item())
            
            loss_history_val_Pretrain = np.append(loss_history_val_Pretrain, loss_val.item())
            loss_u_history_val_Pretrain = np.append(loss_u_history_val_Pretrain, loss_u_val.item())
            loss_f_history_val_Pretrain = np.append(loss_f_history_val_Pretrain, loss_f_val.item())
            
            lambda_history_Pretrain = np.append(lambda_history_Pretrain, lamu.detach().cpu().numpy(), axis=1)

    def adaptive_adjust_collocation_points(self, name_list, project_root, collection_multiple, directory_path, log_flag=False):
        """
        Adaptively adjust collocation points based on derivative magnitude
        
        Parameters:
        -----------
        name_list : list
            List of filenames
        project_root : str
            Root directory of the project
        collection_multiple : int
            Multiplier for collocation point generation
        directory_path : str
            Path to data directory
        log_flag : bool
            Flag for logarithmic scaling
            
        Returns:
        --------
        numpy array
            Adaptively selected collocation points
        """
        directory = project_root
        Data_Collection_Only = np.array([])
        
        for gam_value in name_list:
            # Construct file path
            file_name = f'{gam_value}/post/Tau_Gammat.mat'
            file_path = os.path.join(directory_path, file_name)
            
            # Read .mat file
            try:
                data = loadmat(file_path)
            except FileNotFoundError:
                print(f'File not found: {file_path}')
                continue
            except Exception as e:
                print(f'Error reading {file_path}: {e}')
                continue
            
            # Extract data components
            t, gammat, gamma0, rho, P0, T, tau_xx, ds, epsilon_strain = (
                data['t'], data['gammat'], data['gamma0'], data['rho'], data['P0'],
                data['T'], data['tau_xx'], data['ds'], data['epsilon_strain'])
            
            w = 2 * math.pi / T  # Angular frequency
            
            # Define strain functions
            def gam(t):
                return gamma0/(w)*(1 - np.cos(w*t))
            
            def gam_t(t):
                return gamma0 * np.sin(w*t)
            
            def gam_tt(t):
                return gamma0 * w * np.cos(w*t)
            
            t_total = t.flatten()
            
            # Prepare strain information
            Data_i = np.zeros((len(t_total), 4))
            Data_i[:, 0] = gam(t_total)
            Data_i[:, 1] = gam_t(t_total)
            Data_i[:, 2] = gam_tt(t_total)
            Data_i[:, 3] = tau_xx.flatten() / P0
            
            # Compute derivatives for adaptive sampling
            G_each = torch.tensor(Data_i[:, 1:4], dtype=torch.float32, device=device, requires_grad=True)
            gamma0_tensor = torch.tensor(gamma0, dtype=torch.float32, device=device).repeat(t.shape[0], 1)
            t_tensor = torch.tensor(t, dtype=torch.float32, device=device, requires_grad=True)
            
            f_adaptive, _, u_t_adaptive = self.net_f(gamma0_tensor, t_tensor, G_each)
            
            # Adaptive sampling based on derivative magnitude
            t_vals = t.flatten()
            ut_abs = np.abs(f_adaptive.detach().cpu().numpy().flatten())
            kde = gaussian_kde(t_vals, weights=ut_abs)
            
            # Generate samples
            num_samples = collection_multiple * (len(t) - 1) + 1
            t_samples = kde.resample(num_samples)[0]
            
            # Clip and sort samples
            t_min = int(min(t_vals))
            t_max = int(max(t_vals))
            t_collection = np.sort(np.clip(t_samples, t_min, t_max))
            t_collection = np.unique(t_collection)
            
            # Prepare collocation data
            data_derivative_i = np.zeros((len(t_collection), 5))
            data_derivative_i[:, 0] = t_collection
            data_derivative_i[:, 1] = gam(t_collection)
            data_derivative_i[:, 2] = gam_t(t_collection)
            data_derivative_i[:, 3] = gam_tt(t_collection)
            
            if log_flag:
                data_derivative_i[:, 4] = np.log(gamma0)
            else:
                data_derivative_i[:, 4] = gamma0
            
            # Append to collection
            if Data_Collection_Only.size == 0:
                Data_Collection_Only = data_derivative_i
            else:
                Data_Collection_Only = np.vstack((Data_Collection_Only, data_derivative_i))
        
        return Data_Collection_Only

    def train(self, nIter):
        """
        Train the model using a combination of pretraining, STRidge regression, and Adam optimization
        
        Parameters:
        -----------
        nIter : int
            Number of ADO (Adaptive Discovery and Optimization) loops
        """
        # Pretraining for good initialization
        print('L-BFGS-B pretraining begins')
        
        def closure_pretrain():
            """Closure function for pretraining optimization"""
            self.optimizer_Pretrain.zero_grad()
            loss, loss_u, loss_f, loss_lambda, loss_val, loss_u_val, loss_f_val = self.loss_function()
            loss.backward()
            self.callback_Pretrain(loss, loss_u, loss_f, loss_lambda, loss_val, loss_u_val, loss_f_val, self.lambda1)
            return loss
        
        self.optimizer_Pretrain.step(closure_pretrain)
        
        # Closure function for Adam optimization
        def closure_adam():
            self.optimizer_Adam.zero_grad()
            loss, _, _, _, _, _, _ = self.loss_function()
            loss.backward()
            return loss
        
        # Increase physics loss coefficient after pretraining
        self.loss_f_coeff = 5
        
        # Adaptive Discovery and Optimization (ADO) loop
        for self.it in range(nIter):
            print(f'ADO iteration: {self.it + 1}')
            
            # Sparse identification using STRidge
            print('STRidge optimization begins')
            self.callTrainSTRidge()
            
            # Adaptively adjust collocation points
            print('Adaptive collocation point adjustment')
            Data_Collection_adaptive = self.adaptive_adjust_collocation_points(
                name_list, project_root, collection_multiple, directory_path, log_flag=False)
            
            # Update collocation points
            self.x_f = torch.tensor(Data_Collection_adaptive[:, 4:5], dtype=torch.float32, device=device, requires_grad=True)
            self.t_f = torch.tensor(Data_Collection_adaptive[:, 0:1], dtype=torch.float32, device=device, requires_grad=True)
            self.G_f = torch.tensor(Data_Collection_adaptive[:, 1:5], dtype=torch.float32, device=device, requires_grad=True)
            
            # Adam optimization
            print('Adam optimization begins')
            start_time = time.time()
            for it_Adam in range(1000):
                self.optimizer_Adam.step(closure_adam)
                
                if it_Adam % 10 == 0:
                    elapsed = time.time() - start_time
                    loss = closure_adam().item()
                    loss_u = torch.mean((self.u - self.net_u(self.x, self.t)) ** 2).item()
                    loss_f = self.loss_f_coeff * torch.mean(self.net_f(self.x_f, self.t_f, self.G_f)[0] ** 2).item()
                    loss_lambda = (1e-7 * torch.norm(self.lambda1, p=1)).item()
                    lambda1_value = self.lambda1.detach().cpu().numpy()
                    
                    print(f'Adam Iteration: {it_Adam}, Log Loss: {loss:.3e}, loss_u: {loss_u:.2e}, '
                          f'loss_f: {loss_f:.2e}, loss_lambda: {loss_lambda:.2e}, Time: {elapsed:.2f}')
                    
                    # Update global loss histories
                    global loss_history_Adam, lambda_history_Adam, loss_u_history_Adam
                    global loss_f_history_Adam, loss_lambda_history_Adam, loss_history_Adam_val
                    global loss_u_history_Adam_val, loss_f_history_Adam_val
                    
                    loss_history_Adam = np.append(loss_history_Adam, loss)
                    lambda_history_Adam = np.append(lambda_history_Adam, lambda1_value, axis=1)
                    loss_u_history_Adam = np.append(loss_u_history_Adam, loss_u)
                    loss_f_history_Adam = np.append(loss_f_history_Adam, loss_f)
                    loss_lambda_history_Adam = np.append(loss_lambda_history_Adam, loss_lambda)
                    
                    loss_history_Adam_val = np.append(loss_history_Adam_val, loss)
                    loss_u_history_Adam_val = np.append(loss_u_history_Adam_val, loss_u)
                    loss_f_history_Adam_val = np.append(loss_f_history_Adam_val, loss_f)
                    
                    lamu = self.lambda1.detach().cpu().numpy()
                    lambda_history_Adam = np.append(lambda_history_Adam, lamu, axis=1)
                    
                    start_time = time.time()
        
        # Final STRidge optimization
        print('Final STRidge optimization')
        self.callTrainSTRidge()

    def predict(self, X_star):
        """
        Make predictions using the trained model
        
        Parameters:
        -----------
        X_star : numpy array
            Input points for prediction with shape (n_samples, 2) containing [x, t]
            
        Returns:
        --------
        numpy array
            Predicted values
        """
        x_star = torch.tensor(X_star[:, 0:1], dtype=torch.float32, device=device)
        t_star = torch.tensor(X_star[:, 1:2], dtype=torch.float32, device=device)
        
        u_star = self.net_u(x_star, t_star).detach().cpu().numpy()
        return u_star

    def callTrainSTRidge(self):
        """Wrapper function for STRidge regression to update SINDy coefficients"""
        lam = 1e-5
        d_tol = 20
        maxit = 100
        STR_iters = 10
        
        l0_penalty = None
        normalize = 2
        split = 0.8
        print_best_tol = False
        
        # Get current SINDy library and target
        f_pred, Phi_pred, u_t_pred = self.net_f(self.x_f, self.t_f, self.G_f)
        
        # Update coefficients using STRidge
        lambda2 = self.TrainSTRidge(Phi_pred.detach().cpu().numpy(), u_t_pred.detach().cpu().numpy(), 
                                   lam, d_tol, maxit, STR_iters, l0_penalty, normalize, split, print_best_tol)
        
        # Update model coefficients
        self.lambda1.data = torch.tensor(lambda2, dtype=torch.float32, device=device)

    def TrainSTRidge(self, R0, Ut, lam, d_tol, maxit, STR_iters=10, l0_penalty=None, normalize=2, split=0.8, print_best_tol=False):
        """
        Train model using STRidge (Sparse Trust Region Ridge Regression)
        
        This implementation is inspired by:
        Rudy, Samuel H., et al. "Data-driven discovery of partial differential equations."
        Science Advances 3.4 (2017): e1602614.
        
        Parameters:
        -----------
        R0 : numpy array
            Feature matrix (SINDy library)
        Ut : numpy array
            Target vector (derivatives)
        lam : float
            Ridge regression regularization parameter
        d_tol : float
            Tolerance for coefficient thresholding
        maxit : int
            Maximum number of iterations
        STR_iters : int
            Number of STRidge iterations
        l0_penalty : float
            Penalty for non-zero coefficients
        normalize : int
            Normalization type
        split : float
            Train-test split ratio
        print_best_tol : bool
            Whether to print the best tolerance
            
        Returns:
        --------
        numpy array
            Optimized sparse coefficients
        """
        # Normalize data
        n, d = R0.shape
        R = np.zeros((n, d), dtype=np.float32)
        
        if normalize != 0:
            Mreg = np.zeros((d, 1))
            for i in range(d):
                Mreg[i] = 1.0 / (np.linalg.norm(R0[:, i], normalize))
                R[:, i] = Mreg[i] * R0[:, i]
            normalize_inner = 0  # Already normalized, no need for inner normalization
        else:
            R = R0
            Mreg = np.ones((d, 1)) * d
            normalize_inner = 2
        
        # Update global history
        global lambda_normalized_history_STRidge
        lambda_normalized_history_STRidge = np.append(lambda_normalized_history_STRidge, Mreg, axis=1)
        
        # Split data into training and testing sets
        np.random.seed(0)  # For reproducibility
        n, _ = R.shape
        train = np.random.choice(n, int(n * split), replace=False)
        test = [i for i in np.arange(n) if i not in train]
        
        TrainR, TestR = R[train, :], R[test, :]
        TrainY, TestY = Ut[train, :], Ut[test, :]
        
        # Initialize tolerance
        d_tol = float(d_tol)
        if self.it == 0:
            self.tol = d_tol
        
        # Initialize with current coefficients
        w_best = self.lambda1.detach().cpu().numpy() / Mreg
        
        # Calculate initial error
        err_f = np.mean((TestY - TestR.dot(w_best)) ** 2)
        
        # Set L0 penalty if not specified
        if l0_penalty is None and self.it == 0:
            self.l0_penalty_0 = err_f
            l0_penalty = self.l0_penalty_0
        elif l0_penalty is None:
            l0_penalty = self.l0_penalty_0
        
        # Calculate initial total error
        err_lambda = l0_penalty * np.count_nonzero(w_best)
        err_best = err_f + err_lambda
        tol_best = 0
        
        # Update global loss histories
        global loss_history_STRidge, loss_f_history_STRidge, loss_lambda_history_STRidge, tol_history_STRidge
        loss_history_STRidge = np.append(loss_history_STRidge, err_best)
        loss_f_history_STRidge = np.append(loss_f_history_STRidge, err_f)
        loss_lambda_history_STRidge = np.append(loss_lambda_history_STRidge, err_lambda)
        tol_history_STRidge = np.append(tol_history_STRidge, tol_best)
        
        # Optimize tolerance
        for iter in range(maxit):
            # Get coefficients with current tolerance
            w = self.STRidge(TrainR, TrainY, lam, STR_iters, self.tol, Mreg, normalize=normalize_inner)
            
            # Calculate error
            err_f = np.mean((TestY - TestR.dot(w)) ** 2)
            err_lambda = l0_penalty * np.count_nonzero(w)
            err = err_f + err_lambda
            
            # Update best solution if current is better
            if err <= err_best:
                err_best = err
                w_best = w
                tol_best = self.tol
                self.tol += d_tol
                
                # Update histories
                loss_history_STRidge = np.append(loss_history_STRidge, err_best)
                loss_f_history_STRidge = np.append(loss_f_history_STRidge, err_f)
                loss_lambda_history_STRidge = np.append(loss_lambda_history_STRidge, err_lambda)
                tol_history_STRidge = np.append(tol_history_STRidge, tol_best)
            else:
                # Reduce tolerance
                self.tol = max([0, self.tol - 2 * d_tol])
                d_tol /= 1.618
                self.tol += d_tol
        
        if print_best_tol:
            print(f"Optimal tolerance: {tol_best}")
        
        # Update optimal tolerance history
        global optimaltol_history
        optimaltol_history = np.append(optimaltol_history, tol_best)
        
        return np.real(np.multiply(Mreg, w_best))

    def STRidge(self, X0, y, lam, maxit, tol, Mreg, normalize=2, print_results=False):
        """
        Sparse Trust Region Ridge Regression (STRidge) algorithm
        
        Parameters:
        -----------
        X0 : numpy array
            Feature matrix
        y : numpy array
            Target vector
        lam : float
            Ridge regularization parameter
        maxit : int
            Maximum iterations
        tol : float
            Coefficient threshold
        Mreg : numpy array
            Normalization factors
        normalize : int
            Normalization type
        print_results : bool
            Whether to print results
            
        Returns:
        --------
        numpy array
            Sparse coefficients
        """
        n, d = X0.shape
        X = np.zeros((n, d), dtype=np.complex64)
        
        # Normalize data if required
        if normalize != 0:
            Mreg = np.zeros((d, 1))
            for i in range(d):
                Mreg[i] = 1.0 / (np.linalg.norm(X0[:, i], normalize))
                X[:, i] = Mreg[i] * X0[:, i]
        else:
            X = X0
        
        # Initialize with current coefficients
        w = self.lambda1.detach().cpu().numpy() / Mreg
        
        # Track large coefficients
        biginds = np.where(abs(w) > tol)[0]
        num_relevant = d
        
        # Update global history
        global lambda_history_STRidge, ridge_append_counter_STRidge
        lambda_history_STRidge = np.append(lambda_history_STRidge, np.multiply(Mreg, w), axis=1)
        ridge_append_counter = 1
        
        # Iterative thresholding
        for j in range(maxit):
            # Identify small coefficients to threshold
            smallinds = np.where(abs(w) < tol)[0]
            new_biginds = [i for i in range(d) if i not in smallinds]
            
            # Check if convergence reached
            if num_relevant == len(new_biginds):
                break
            num_relevant = len(new_biginds)
            
            # Handle case with no relevant features
            if len(new_biginds) == 0:
                if j == 0:
                    if normalize != 0:
                        lambda_history_STRidge = np.append(lambda_history_STRidge, w * Mreg, axis=1)
                    else:
                        lambda_history_STRidge = np.append(lambda_history_STRidge, w * Mreg, axis=1)
                    ridge_append_counter_STRidge = np.append(ridge_append_counter_STRidge, ridge_append_counter)
                    return np.multiply(Mreg, w) if normalize != 0 else w
                else:
                    break
            
            # Update indices and set small coefficients to zero
            biginds = new_biginds
            w[smallinds] = 0
            
            # Solve ridge regression for remaining coefficients
            if lam != 0:
                w[biginds] = np.linalg.lstsq(
                    X[:, biginds].T.dot(X[:, biginds]) + lam * np.eye(len(biginds)),
                    X[:, biginds].T.dot(y)
                )[0]
            else:
                w[biginds] = np.linalg.lstsq(X[:, biginds], y)[0]
            
            # Update history
            lambda_history_STRidge = np.append(lambda_history_STRidge, np.multiply(Mreg, w), axis=1)
            ridge_append_counter += 1
        
        # Final least squares fit with identified sparsity pattern
        if biginds:
            w[biginds] = np.linalg.lstsq(X[:, biginds], y)[0]
        
        # Update history and return
        lambda_history_STRidge = np.append(lambda_history_STRidge, w * Mreg, axis=1)
        ridge_append_counter += 1
        ridge_append_counter_STRidge = np.append(ridge_append_counter_STRidge, ridge_append_counter)
        
        return np.multiply(Mreg, w) if normalize != 0 else w


# Main execution
if __name__ == "__main__":
    start_time = time.time()
    layers = [2, 20, 20, 20, 20, 20, 20, 1]
    collection_multiple = 1
    
    # Set up paths
    project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    directory_path = os.path.join(
        project_root, 'Datasets', 'Representative_Configuration', 'Sinusoidal', '16_groups'
    )
    today_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    
    # Initialize data handlers
    GD = GenerateData()
    reader = ReadFolder()
    classification_results = reader.return_dimensionless_cases(directory_path)
    store_excel_path = f"{directory_path}/Results_{today_datetime}"
    
    # Process each classification case
    for item in classification_results:
        try:
            del model
            print("Model has been deleted.")
        except NameError:
            print("Model does not exist, nothing to delete.")
        
        # Reset loss histories for each case
        loss_history_Adam = np.array([0])
        loss_u_history_Adam = np.array([0])
        loss_f_history_Adam = np.array([0])
        loss_lambda_history_Adam = np.array([0])
        lambda_history_Adam = np.zeros((num_coe, 1))
        loss_history_Adam_val = np.array([0])
        loss_u_history_Adam_val = np.array([0])
        loss_f_history_Adam_val = np.array([0])
        
        loss_history_STRidge = np.array([0])
        loss_f_history_STRidge = np.array([0])
        loss_lambda_history_STRidge = np.array([0])
        optimaltol_history = np.array([0])
        tol_history_STRidge = np.array([0])
        lambda_normalized_history_STRidge = np.zeros((num_coe, 1))
        
        lambda_history_STRidge = np.zeros((num_coe, 1))
        ridge_append_counter_STRidge = np.array([0])
        
        loss_history_Pretrain = np.array([0])
        loss_u_history_Pretrain = np.array([0])
        loss_f_history_Pretrain = np.array([0])
        loss_lambda_history_Pretrain = np.array([0])
        loss_history_val_Pretrain = np.array([0])
        loss_u_history_val_Pretrain = np.array([0])
        loss_f_history_val_Pretrain = np.array([0])
        step_Pretrain = 0
        
        lambda_history_Pretrain = np.zeros((num_coe, 1))
        
        # Reset random seeds
        np.random.seed(1234)
        torch.manual_seed(1234)
        torch.cuda.manual_seed_all(42)
        
        # Set up storage path
        store_path = f"{directory_path}/Results_{today_datetime}/{item['category']}"
        if not os.path.exists(store_path):
            os.makedirs(store_path)
            print(f"Directory created: '{store_path}'")
        else:
            print(f"Directory already exists: '{store_path}'")
        
        # Prepare data
        name_list = item['folders']
        Data_Total, Data_Train, Data_Validation, Data_Collection, lb, ub, t = \
            GD.create_DimensionalCase_data(name_list, project_root, collection_multiple, directory_path, log_flag=False)
        
        # Initialize and train model
        model = PINNSR(Data_Train, Data_Collection, Data_Validation, layers, lb, ub,
                                 name_list, project_root, collection_multiple, directory_path, log_flag=False)
        model.train(10)  # Perform 10 ADO iterations
        
        # Validation and visualization
        X_u_meas = np.vstack((Data_Train[:, 5], Data_Train[:, 0])).T
        u_meas = Data_Train[:, 4].reshape(-1, 1)
        
        # Extract variables
        t_total = Data_Total[:, 0]
        γ = Data_Total[:, 1]
        γ_t = Data_Total[:, 2]
        γ_tt = Data_Total[:, 3]
        σ = Data_Total[:, 4]
        abs_γ_t = abs(γ_t)
        abs_γ_tt = abs(γ_tt)
        
        # Plot training data
        plt.figure()
        plt.scatter(X_u_meas[:,1], u_meas[:,0])
        plt.scatter(t_total, σ)
        
        # Plot validation results
        xx = torch.tensor(Data_Validation[:, 5:6], dtype=torch.float32, device=device, requires_grad=True)
        tt = torch.tensor(Data_Validation[:, 0:1], dtype=torch.float32, device=device, requires_grad=True)
        u_validation_pred = model.net_u(xx, tt)
        u_validation_true = Data_Validation[:, 4:5]
        
        # Configure plot style
        mpl.rcParams['font.family'] = 'Times New Roman'
        mpl.rcParams['font.serif'] = 'Times New Roman'
        
        plt.figure(figsize=(10, 6))
        plt.scatter(tt.detach().cpu().numpy(), u_validation_true, label='σ_DEM', color='blue', s=10, alpha=0.99)
        plt.scatter(tt.detach().cpu().numpy(), u_validation_pred.detach().cpu().numpy(), 
                   label='σ_NN', color='red', marker='x', s=10, alpha=0.6)
        plt.xlabel('t (s)', fontsize=18)
        plt.ylabel(r'$\sigma_{xy}$/$P_0$', fontsize=18)
        plt.tick_params(axis='both', which='major', labelsize=12)
        plt.legend(fontsize=14)
        plt.savefig(f'{store_path}/NN_Predicted_Result.png')
        
        # Plot time derivatives
        xx0 = torch.tensor(Data_Total[:, 5:6], dtype=torch.float32, device=device, requires_grad=True)
        tt0 = torch.tensor(Data_Total[:, 0:1], dtype=torch.float32, device=device, requires_grad=True)
        u_total_pred0 = model.net_u(xx0, tt0)
        u_total_grad0 = torch.autograd.grad(u_total_pred0, tt0, grad_outputs=torch.ones_like(u_total_pred0), create_graph=True)[0]
        
        plt.figure()
        plt.scatter(tt0.detach().cpu().numpy(), u_total_grad0.detach().cpu().numpy(), 
                   label='σ_t calculated by AD', color='blue', s=10)
        plt.xlabel('t (s)')
        plt.ylabel(r'$\partial \sigma_{xy} / \partial t$')
        plt.legend()
        
        # Plot SINDy equation results
        xx1 = torch.tensor(Data_Total[:, 5:6], dtype=torch.float32, device=device, requires_grad=True)
        tt1 = torch.tensor(Data_Total[:, 0:1], dtype=torch.float32, device=device, requires_grad=True)
        u_total_pred = model.net_u(xx1, tt1)
        u_total_grad = torch.autograd.grad(u_total_pred, tt1, grad_outputs=torch.ones_like(u_total_pred), create_graph=True)[0]
        
        plt.figure()
        plt.plot(tt1.detach().cpu().numpy(), u_total_grad.detach().cpu().numpy(), 
                label='σ_t calculated by AD', color='blue', linewidth=1.5)
        plt.xlabel('t (s)')
        plt.ylabel(r'$\partial \sigma_{xy} / \partial t$')
        plt.legend()
        
        # Verify SINDy equation
        σ_t = u_total_grad.squeeze(dim=1).detach().cpu().numpy()
        ones_tensor = np.ones((t_total.shape[0], ))
        
        # Construct feature matrix
        features = [ones_tensor, σ, γ_t, γ_tt, σ*σ, γ_t*γ_t, γ_tt*γ_tt, abs_γ_t, abs_γ_tt,
                   σ*γ_t, σ*γ_tt, σ*abs_γ_t, σ*abs_γ_tt, abs_γ_t*γ_t, abs_γ_t*γ_tt,
                   (σ*σ)*γ_t, (σ*σ)*γ_tt, (σ*σ)*abs_γ_t, (σ*σ)*abs_γ_tt, γ_t*(γ_tt*γ_tt), 
                   γ_tt*(γ_t*γ_t), σ_t*γ_t, σ_t*abs_γ_t]
        
        # Get SINDy coefficients
        lambda1_value = model.lambda1.detach().cpu().numpy()
        lambda1_true = np.zeros((num_coe, 1))
        lambda1_true = np.ravel(lambda1_true)
        lambda1_value = np.ravel(lambda1_value)
        
        # Predict using SINDy equation
        Phi = np.vstack(features).T
        σ_t_pred = np.dot(Phi, lambda1_value)
        
        # Plot comparison
        plt.figure(figsize=(10, 6))
        plt.scatter(tt1.detach().cpu().numpy(), u_total_grad.detach().cpu().numpy(), 
                   label=r'$\mu_t$_AutoDiff', color='blue', s=10)
        plt.scatter(tt1.detach().cpu().numpy(), σ_t_pred, 
                   label=r'$\mu_t$_Equation', color='red', marker='x', s=10, alpha=0.6)
        plt.xlabel('t (s)')
        plt.ylabel(r'$\mu_t$')
        plt.legend()
        plt.savefig(f'{store_path}/σ_t_Predicted_Result.png')
        
        # Calculate and save errors
        store_path_full = f'{store_path}/'
        f = open(f'{store_path_full}ErrorRecord.txt', "a+")
        
        # Prepare data for error calculation
        X_u_train = torch.tensor(np.concatenate((Data_Train[:, 1:2], Data_Train[:, 0:1]), axis=1))
        u_train = torch.tensor(np.concatenate((Data_Train[:, 4:5], Data_Train[:, 0:1]), axis=1))
        
        X_u_val = torch.tensor(np.concatenate((Data_Validation[:, 1:2], Data_Validation[:, 0:1]), axis=1))
        u_val = torch.tensor(np.concatenate((Data_Validation[:, 4:5], Data_Validation[:, 0:1]), axis=1))
        
        # Calculate errors
        u_train_Pred = model.predict(X_u_train)
        Error_u_Train = np.linalg.norm(u_train - u_train_Pred, 2) / np.linalg.norm(u_train, 2)
        f.write(f'Training Error u: {Error_u_Train:.2e} \n')
        
        u_val_Pred = model.predict(X_u_val)
        Error_u_Val = np.linalg.norm(u_val - u_val_Pred, 2) / np.linalg.norm(u_val, 2)
        f.write(f'Validation Error u: {Error_u_Val:.2e} \n')
        
        elapsed = time.time() - start_time
        f.write(f'Training time: {elapsed:.4f} \n')
        f.close()
        
        with open(f'{store_path_full}Output.txt', 'w') as f:
            # Write discovered equation
            disc_eq_temp = []
            for i_lib in range(len(model.library_description)):
                if lambda1_value[i_lib] != 0:
                    disc_eq_temp.append(f'{lambda1_value[i_lib]:.4f}{model.library_description[i_lib]}')
            disc_eq = '+'.join(disc_eq_temp)
            f.write(f'The discovered equation: σ_t = {disc_eq}')
        
        # Plot identified coefficients
        plt.figure()
        plt.plot(lambda1_value, label='the identified')
        plt.title('lambda values')
        plt.legend()
        plt.savefig(f'{store_path_full}identified_coeff.png')
        
        # Save coefficients to Excel
        excel_path = f'{store_excel_path}/SINDy_Coefficients.xlsx'
        library_descriptions = model.library_description
        lambda_values = lambda1_value
        
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
            ws = wb.active
        else:
            df = pd.DataFrame([library_descriptions], index=['Library Description'])
            df.to_excel(excel_path, index=True, engine='openpyxl')
            wb = load_workbook(excel_path)
            ws = wb.active
        
        current_row = ws.max_row
        lambda_values_list = lambda_values.tolist()
        index_name = f"{item['category']}"
        lambda_values_list.insert(0, index_name)
        ws.append(lambda_values_list)
        
        # Format cells
        number_format = NamedStyle(name='number_format', number_format='0.000')
        alignment = Alignment(horizontal='center', vertical='center')
        for row in ws.iter_rows(min_row=current_row + 1, max_row=current_row + 1):
            for cell in row:
                cell.alignment = alignment
        
        # Adjust column widths
        def adjust_column_width(ws):
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                ws.column_dimensions[column].width = adjusted_width
        
        adjust_column_width(ws)
        wb.save(excel_path)
        
        # Save results for error analysis
        end_time_total = time.time()
        run_time = end_time_total - start_time
        
        variables = {
            'lambda_history_Adam': lambda_history_Adam,
            'lambda_history_Pretrain': lambda_history_Pretrain,
            'lambda_history_STRidge': lambda_history_STRidge,
            'lambda_normalized_history_STRidge': lambda_normalized_history_STRidge,
            'lambda_values': lambda_values,
            'optimaltol_history': optimaltol_history,
            'tol_history_STRidge': tol_history_STRidge,
            'run_time': run_time
        }
        
        # Add all loss variables
        local_vars = list(locals().keys())
        for var_name in local_vars:
            if var_name.startswith('loss'):
                variables[var_name] = locals()[var_name]
        
        scipy.io.savemat(f'{store_path_full}ErrorAnalysis.mat', variables)
        
        del model
