# -*- coding: utf-8 -*-
"""
PINNSR Framework for case-specific governing equation discovery.

======================================================================================================================
Created on Mon Jun 24 20:32:37 2024

Inspired by:
1. Chen, Z., Liu, Y., & Sun, H. (2021). Physics-informed learning of governing equations from scarce data. 
   Nature Communications, 12(1), 6136.
2. Rudy, S. H., Brunton, S. L., Proctor, J. L., & Kutz, J. N. (2017). Data-driven discovery of partial 
   differential equations. Science Advances, 3(4), e1602614.       

Novel Contributions:
1. Enhanced PINN-SR architecture with input-output structures tailored for granular rheological conditions.
2. Adaptive collocation sampling strategy to prioritize physical information in high-gradient regions 
   induced by shear reversal.
======================================================================================================================

@author: Han Xu
"""

#%% Import required libraries
import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
import sympy as sp  # For symbolic differentiation
import autograd.numpy as agnp
from autograd import grad, elementwise_grad
from scipy.io import loadmat

import torch
import torch.nn as nn
import scipy.io
from scipy.interpolate import griddata
from scipy.spatial import distance
from matplotlib import cm
import matplotlib as mpl
import time
from mpl_toolkits.mplot3d import Axes3D
from pyDOE import lhs
from torch.autograd.functional import jacobian
import math
from pysindy.differentiation import FiniteDifference, SmoothedFiniteDifference
import openpyxl
from openpyxl import load_workbook, Workbook
import sys
from scipy.stats import gaussian_kde
from datetime import datetime

# Set up project paths
project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
src_path = os.path.join(project_root, 'src')
sys.path.append(src_path)

# Import custom modules
from GenerateData_AdaptionCollection import GenerateData
from ReadFolder import ReadFolder

# Set device for computation (GPU if available)
device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')

# Record total execution time
start_time_total = time.time()

#%% Configuration parameters
num_coe = 23  # Number of candidate terms in SINDy library

# =============================================================================
# Define loss history arrays to record convergence
# =============================================================================
# Adam optimizer loss history
loss_history_Adam = np.array([0])
loss_u_history_Adam = np.array([0])
loss_f_history_Adam = np.array([0])
loss_lambda_history_Adam = np.array([0])
lambda_history_Adam = np.zeros((num_coe, 1))

# Import Excel styling modules
from openpyxl.styles import Alignment, NamedStyle

# Validation loss history for Adam
loss_history_Adam_val = np.array([0])
loss_u_history_Adam_val = np.array([0])
loss_f_history_Adam_val = np.array([0])

# STRidge optimizer loss history
loss_history_STRidge = np.array([0])
loss_f_history_STRidge = np.array([0])
loss_lambda_history_STRidge = np.array([0])
optimaltol_history = np.array([0])
tol_history_STRidge = np.array([0])
lambda_normalized_history_STRidge = np.zeros((num_coe, 1))

lambda_history_STRidge = np.zeros((num_coe, 1))
ridge_append_counter_STRidge = np.array([0])

# Loss histories for pretraining
loss_history_Pretrain = np.array([0])
loss_u_history_Pretrain = np.array([0])
loss_f_history_Pretrain = np.array([0])
loss_lambda_history_Pretrain = np.array([0])
loss_history_val_Pretrain = np.array([0])
loss_u_history_val_Pretrain = np.array([0])
loss_f_history_val_Pretrain = np.array([0])
step_Pretrain = 0

lambda_history_Pretrain = np.zeros((num_coe, 1))

# Set random seeds for reproducibility
np.random.seed(1234)
torch.manual_seed(1234)
torch.cuda.manual_seed_all(42)

class PhysicsInformedNN:
    """
    Physics-Informed Neural Network (PINN) combined with SINDy for system identification.
    
    This class implements a neural network that incorporates physical constraints
    through a modified loss function, and uses the SINDy algorithm to identify
    sparse nonlinear dynamical systems from data.
    """
    
    def __init__(self, Data_Train, Data_Collection, Data_Validation, layers, lb, ub, 
                 name_list, project_root, collection_multiple, directory_path, h_alpha, Pre_step, dtol, log_flag=False):
        """
        Initialize the PhysicsInformedNN model.
        
        Parameters:
        -----------
        Data_Train : numpy array
            Training dataset with shape (n_samples, 5) containing [t, γ, γ_t, γ_tt, σ]
        Data_Collection : numpy array
            Collocation points with shape (n_points, 4) containing [t, γ, γ_t, γ_tt]
        Data_Validation : numpy array
            Validation dataset with shape (n_samples, 5) containing [t, γ, γ_t, γ_tt, σ]
        layers : list
            Neural network architecture (number of neurons per layer)
        lb : numpy array
            Lower bounds of input space [t_min, γ_min]
        ub : numpy array
            Upper bounds of input space [t_max, γ_max]
        name_list : list
            List of dataset names
        project_root : str
            Root directory of the project
        collection_multiple : int
            Multiplier for collocation point density
        directory_path : str
            Path to data directory
        h_alpha : float
            Coefficient for physics loss term
        Pre_step : int
            Number of pre-training steps
        dtol : float
            Tolerance for SINDy coefficient thresholding
        log_flag : bool
            Whether to use logarithmic transformation on inputs
        """
        self.lb = torch.tensor(lb, dtype=torch.float32, device=device)
        self.ub = torch.tensor(ub, dtype=torch.float32, device=device)
        self.layers = layers
        self.it = 0  # Iteration counter for ADO loop
        
        # Initialize neural network weights and biases
        self.weights, self.biases = self.initialize_NN(layers)
        
        # Initialize SINDy coefficients as trainable parameters
        self.lambda1 = nn.Parameter(torch.zeros(num_coe, 1, dtype=torch.float32, device=device, requires_grad=True))
        
        # Prepare training data tensors
        self.x = torch.tensor(Data_Train[:, 5:6], dtype=torch.float32, device=device, requires_grad=True)  # gamma0
        self.t = torch.tensor(Data_Train[:, 0:1], dtype=torch.float32, device=device, requires_grad=True)  # time
        self.u = torch.tensor(Data_Train[:, 4:5], dtype=torch.float32, device=device, requires_grad=True)  # sigma (output)
        self.x_f = torch.tensor(Data_Collection[:, 4:5], dtype=torch.float32, device=device, requires_grad=True)  # collocation points gamma0
        self.t_f = torch.tensor(Data_Collection[:, 0:1], dtype=torch.float32, device=device, requires_grad=True)  # collocation points time
        self.x_val = torch.tensor(Data_Validation[:, 5:6], dtype=torch.float32, device=device, requires_grad=True)  # validation gamma0
        self.t_val = torch.tensor(Data_Validation[:, 0:1], dtype=torch.float32, device=device, requires_grad=True)  # validation time
        self.u_val = torch.tensor(Data_Validation[:, 4:5], dtype=torch.float32, device=device, requires_grad=True)  # validation sigma
        
        # Strain information matrices
        self.G_f = torch.tensor(Data_Collection[:, 1:5], dtype=torch.float32, device=device, requires_grad=True)  # collocation gamma information
        self.G_val = torch.tensor(Data_Validation[:, 1:5], dtype=torch.float32, device=device, requires_grad=True)  # validation gamma information
        
        # Collect all trainable parameters
        self.parameters = self.weights + self.biases + [self.lambda1]
        
        # Define optimizers
        var_list_1 = self.biases + self.weights
        self.optimizer_Adam = torch.optim.Adam(var_list_1, lr=1e-3, betas=(0.99, 0.9), eps=1e-8)
        self.optimizer_Pretrain = torch.optim.LBFGS(self.parameters, max_iter=Pre_step, history_size=50, 
                                                   max_eval=Pre_step, tolerance_grad=1e-7, 
                                                   tolerance_change=1e-9, line_search_fn="strong_wolfe")
        self.optimizer_Post = torch.optim.LBFGS(self.parameters, max_iter=20000, history_size=50, 
                                               max_eval=20000, tolerance_grad=1e-7, 
                                               tolerance_change=1e-9, line_search_fn="strong_wolfe")
        self.loss_f_coeff = 0.01  # Initial coefficient for physics loss term
    
    def initialize_NN(self, layers):
        """Initialize neural network weights using Xavier initialization."""
        weights = []
        biases = []
        num_layers = len(layers)
        for l in range(num_layers - 1):
            W = self.xavier_init(size=[layers[l], layers[l + 1]])
            b = torch.zeros([1, layers[l + 1]], dtype=torch.float32, device=device, requires_grad=True)
            weights.append(W)
            biases.append(b)
        return weights, biases
    
    def xavier_init(self, size):
        """Xavier initialization for neural network weights."""
        in_dim = size[0]
        out_dim = size[1]
        xavier_stddev = np.sqrt(2 / (in_dim + out_dim))
        return torch.nn.Parameter(torch.randn([in_dim, out_dim], dtype=torch.float32, device=device) * xavier_stddev)

    def neural_net(self, x, t, weights, biases):
        """Forward pass through the neural network."""
        X = torch.cat((x, t), dim=1)
        num_layers = len(weights) + 1
        # Normalize input to [-1, 1] range
        H = 2.0 * (X - self.lb) / (self.ub - self.lb) - 1.0
        for l in range(num_layers - 2):
            W = weights[l]
            b = biases[l]
            H = torch.tanh(torch.add(torch.matmul(H, W), b))
        # Final layer without activation
        W = weights[-1]
        b = biases[-1]
        Y = torch.add(torch.matmul(H, W), b)
        return Y
    
    def net_u(self, x, t):
        """Predict the output variable (sigma) using the neural network."""
        u = self.neural_net(x, t, self.weights, self.biases)
        return u
    
    def net_f(self, x, t, G):
        """
        Compute the residual of the SINDy-predicted dynamics.
        
        Constructs the SINDy library and calculates the discrepancy between
        the neural network's time derivative and the SINDy prediction.
        
        Parameters:
        -----------
        x : torch.Tensor
            Input variable (gamma0)
        t : torch.Tensor
            Time variable
        G : torch.Tensor
            Strain information matrix [γ, γ', γ'']
            
        Returns:
        --------
        f : torch.Tensor
            Residual of the dynamical system prediction
        Phi : torch.Tensor
            SINDy library matrix
        u_t : torch.Tensor
            Time derivative of the output variable
        """
        u = self.net_u(x, t)
        # Compute time derivative using automatic differentiation
        u_t = torch.autograd.grad(u, t, grad_outputs=torch.ones_like(u), create_graph=True)[0]
        
        # Extract strain components from G matrix
        g = G[:, 0:1]
        g_t = G[:, 1:2]
        g_tt = G[:, 2:3]
        abs_g_t = torch.abs(G[:, 1:2])
        abs_g_tt = torch.abs(G[:, 2:3])
        
        # Create constant term tensor
        ones_tensor = torch.ones((x.shape[0], 1), dtype=torch.float32, device=device)

        # Construct SINDy library - candidate terms for the dynamical system
        Phi = torch.cat([ones_tensor, u, g_t, g_tt, u*u, g_t*g_t, g_tt*g_tt, abs_g_t, abs_g_tt,
                         u*g_t, u*g_tt, u*abs_g_t, u*abs_g_tt, abs_g_t*g_t, abs_g_t*g_tt,
                         (u*u)*g_t, (u*u)*g_tt, (u*u)*abs_g_t, (u*u)*abs_g_tt, g_t*(g_tt*g_tt), 
                         g_tt*(g_t*g_t), u_t*g_t, u_t*abs_g_t], 1)
        
        Phi = Phi.to(torch.float32)

        # Library term descriptions for later interpretation
        self.library_description = ['1', 'σ', 'γ_t', 'γ_tt', 'σ^2', 'γ_t^2', 'γ_tt^2', '|γ_t|', '|γ_tt|',
                                    'σ*γ_t', 'σ*γ_tt', 'σ*|γ_t|', 'σ*|γ_tt|', '|γ_t|*γ_t', '|γ_t|*γ_tt',
                                    'σ^2*γ_t', 'σ^2*γ_tt', 'σ^2*|γ_t|', 'σ^2*|γ_tt|', 'γ_t*γ_tt^2', 
                                    'γ_t^2*γ_tt', 'σ_t*γ_t', 'σ_t*|γ_t|']

        # Calculate residual: SINDy prediction - actual derivative
        f = torch.matmul(Phi, self.lambda1) - u_t
        return f, Phi, u_t

    def loss_function(self):
        """
        Calculate the composite loss function.
        
        Combines data loss, physics loss, and regularization terms.
        Also computes validation loss for monitoring generalization.
        """
        # Predictions on training data
        u_pred = self.net_u(self.x, self.t)
        
        # Physics loss components
        f_pred, Phi_pred, u_t_pred = self.net_f(self.x_f, self.t_f, self.G_f)
        
        # Predictions on validation data
        u_val_pred = self.net_u(self.x_val, self.t_val)
        f_val_pred, Phi_val_pred, u_t_val_pred = self.net_f(self.x_val, self.t_val, self.G_val)

        # Calculate individual loss components
        loss_u = torch.mean((self.u - u_pred) ** 2)  # Data mismatch loss
        loss_f = self.loss_f_coeff * torch.mean(f_pred ** 2)  # Physics loss
        loss_lambda = 1e-7 * torch.norm(self.lambda1, p=1)  # L1 regularization on coefficients

        # Validation loss components
        loss_u_val = torch.mean((self.u_val - u_val_pred) ** 2)
        loss_f_val = self.loss_f_coeff * torch.mean(f_val_pred ** 2)

        # Composite loss (logarithmic for better scaling)
        loss = torch.log(loss_u + loss_f + loss_lambda)
        loss_val = torch.log(loss_u_val + loss_f_val)
        return loss, loss_u, loss_f, loss_lambda, loss_val, loss_u_val, loss_f_val

    def callback_Pretrain(self, loss, loss_u, loss_f, loss_lambda, loss_val, loss_u_val, loss_f_val, lamu):
        """Callback function to record training progress during pretraining."""
        global step_Pretrain
        step_Pretrain += 1
        if step_Pretrain % 10 == 0:
            print('Step: %d, log Loss: %e, loss_u: %e, loss_f: %e, loss_lambda: %e' % 
                  (step_Pretrain, loss, loss_u, loss_f, loss_lambda))

            # Update global loss history arrays
            global loss_history_Pretrain, loss_u_history_Pretrain, loss_f_history_Pretrain
            global loss_lambda_history_Pretrain, loss_history_val_Pretrain
            global loss_u_history_val_Pretrain, loss_f_history_val_Pretrain, lambda_history_Pretrain

            loss_history_Pretrain = np.append(loss_history_Pretrain, loss.item())
            loss_u_history_Pretrain = np.append(loss_u_history_Pretrain, loss_u.item())
            loss_f_history_Pretrain = np.append(loss_f_history_Pretrain, loss_f.item())
            loss_lambda_history_Pretrain = np.append(loss_lambda_history_Pretrain, loss_lambda.item())

            loss_history_val_Pretrain = np.append(loss_history_val_Pretrain, loss_val.item())
            loss_u_history_val_Pretrain = np.append(loss_u_history_val_Pretrain, loss_u_val.item())
            loss_f_history_val_Pretrain = np.append(loss_f_history_val_Pretrain, loss_f_val.item())

            lambda_history_Pretrain = np.append(lambda_history_Pretrain, lamu.detach().cpu().numpy(), axis=1)
    
    def callback_Post(self, loss, loss_u, loss_f, loss_lambda, loss_val, loss_u_val, loss_f_val, lamu):
        """Callback function to record training progress during post-training."""
        global step_Pt
        step_Pt += 1
        if step_Pt % 10 == 0:
            print('Step: %d, log Loss: %e, loss_u: %e, loss_f: %e, loss_lambda: %e' % 
                  (step_Pt, loss, loss_u, loss_f, loss_lambda))

            # Update global loss history arrays
            global loss_history_Pt, loss_u_history_Pt, loss_f_history_Pt
            global loss_lambda_history_Pt, loss_history_val_Pt
            global loss_u_history_val_Pt, loss_f_history_val_Pt, lambda_history_Pt

            loss_history_Pt = np.append(loss_history_Pt, loss.item())
            loss_u_history_Pt = np.append(loss_u_history_Pt, loss_u.item())
            loss_f_history_Pt = np.append(loss_f_history_Pt, loss_f.item())
            loss_lambda_history_Pt = np.append(loss_lambda_history_Pt, loss_lambda.item())

            loss_history_val_Pt = np.append(loss_history_val_Pt, loss_val.item())
            loss_u_history_val_Pt = np.append(loss_u_history_val_Pt, loss_u_val.item())
            loss_f_history_val_Pt = np.append(loss_f_history_val_Pt, loss_f_val.item())

            lambda_history_Pt = np.append(lambda_history_Pt, lamu.detach().cpu().numpy(), axis=1)
    
    def adaptive_adjust_collocation_points(self, name_list, project_root, collection_multiple, directory_path, log_flag=False):
        """
        Adaptively adjust collocation points based on residual magnitude.
        
        Creates new collocation points with higher density in regions where
        the model residual is larger, improving accuracy in important regions.
        """
        directory = project_root
        Data_Collection_Only = np.array([])
        
        for gam_value in name_list:
            # Construct file path
            file_name = f'{gam_value}/post/Tau_Gammat.mat'
            file_path = os.path.join(directory_path, file_name)
                
            # Read .mat file
            try:
                data = loadmat(file_path)
            except FileNotFoundError:
                print(f'File not found: {file_path}')
                continue
            except Exception as e:
                print(f'Error reading {file_path}: {e}')
                continue
                
            # Extract data components
            t, gammat, gamma0, rho, P0, T, tau_xx, ds, epsilon_strain = (
                data['t'], data['gammat'], data['gamma0'], data['rho'], data['P0'],
                data['T'], data['tau_xx'], data['ds'], data['epsilon_strain'])
            
            w = 2 * math.pi / T  # Angular frequency
            
            # Define strain functions
            def gam(t):
                return gamma0/(w)*(1 - np.cos(w*t))
            
            def gam_t(t):
                return gamma0 * np.sin(w*t)
            
            def gam_tt(t):
                return gamma0 * w * np.cos(w*t)
            
            t_total = t.flatten()
            
            # Prepare strain data
            Data_i = np.zeros((len(t_total), 4))
            Data_i[:, 0] = gam(t_total)
            Data_i[:, 1] = gam_t(t_total)
            Data_i[:, 2] = gam_tt(t_total)
            Data_i[:, 3] = tau_xx.flatten() / P0
          
            # Calculate current residuals for adaptive sampling
            G_each = torch.tensor(Data_i[:, 1:4], dtype=torch.float32, device=device, requires_grad=True)
            gamma0_tensor = torch.tensor(gamma0, dtype=torch.float32, device=device).repeat(t.shape[0], 1)
            t_tensor = torch.tensor(t, dtype=torch.float32, device=device, requires_grad=True)
            
            f_adaptive, Phi_adaptive, u_t_adaptive = self.net_f(gamma0_tensor, t_tensor, G_each)
            
            # Calculate current predictions and derivatives
            u_current = self.net_u(gamma0_tensor, t_tensor)
            ut_current = torch.autograd.grad(u_current, t_tensor, grad_outputs=torch.ones_like(u_current), create_graph=True)[0]
            
            # Adaptive sampling based on residual magnitude
            t_vals = t.flatten()
            ut_abs_squared = np.square(np.abs(f_adaptive.detach().cpu().numpy().flatten()))
            
            # Use Gaussian kernel density estimation for sampling
            kde = gaussian_kde(t_vals, weights=ut_abs_squared)
            num_samples = collection_multiple * (len(t) - 1) + 1
            t_samples = kde.resample(num_samples)[0]
            
            # Clip and sort samples
            t_min = int(min(t_vals))
            t_max = int(max(t_vals))
            t_collection = np.sort(np.clip(t_samples, t_min, t_max))
            t_collection = np.unique(t_collection)
            
            # Generate new collocation points
            data_derivative_i = np.zeros((len(t_collection), 5))
            data_derivative_i[:, 0] = t_collection
            data_derivative_i[:, 1] = gam(t_collection)
            data_derivative_i[:, 2] = gam_t(t_collection)
            data_derivative_i[:, 3] = gam_tt(t_collection)
            
            if log_flag:
                data_derivative_i[:, 4] = np.log(gamma0)
            else:
                data_derivative_i[:, 4] = gamma0

            # Accumulate collocation points
            if Data_Collection_Only.size:
                Data_Collection_Only = np.vstack((Data_Collection_Only, data_derivative_i))
            else:
                Data_Collection_Only = data_derivative_i
        
        return Data_Collection_Only
    
    def train(self, nIter):
        """
        Main training loop for the PINN-SINDy model.
        
        Implements a multi-stage training process:
        1. Pretraining with L-BFGS
        2. SINDy coefficient identification with STRidge
        3. Fine-tuning with Adam optimizer
        4. Final optimization with L-BFGS
        """
        print('L-BFGS-B pretraining begins')

        # Define closure function for pretraining
        def closure_pretrain():
            self.optimizer_Pretrain.zero_grad()
            loss, loss_u, loss_f, loss_lambda, loss_val, loss_u_val, loss_f_val = self.loss_function()
            loss.backward()
            self.callback_Pretrain(loss, loss_u, loss_f, loss_lambda, loss_val, loss_u_val, loss_f_val, self.lambda1)
            return loss

        self.optimizer_Pretrain.step(closure_pretrain)
                                                       
        # Define closure function for Adam optimization
        def closure_adam():
            self.optimizer_Adam.zero_grad()
            loss, loss_u, loss_f, loss_lambda, loss_val, loss_u_val, loss_f_val = self.loss_function()
            loss.backward()
            return loss

        self.loss_f_coeff = h_alpha  # Update physics loss coefficient
        
        # Main training loop with adaptive discovery and optimization (ADO)
        for self.it in range(nIter):
            print(f'ADO iteration: {self.it + 1}')
            
            # Sparse identification with STRidge
            print('STRidge optimization begins')
            self.callTrainSTRidge()
            
            # Adaptively adjust collocation points based on current model residuals
            print('Adaptive collocation point adjustment')
            Data_Collection_adaptive = self.adaptive_adjust_collocation_points(
                name_list, project_root, collection_multiple, directory_path, log_flag=False)
            
            # Update collocation points in the model
            self.x_f = torch.tensor(Data_Collection_adaptive[:, 4:5], dtype=torch.float32, device=device, requires_grad=True)
            self.t_f = torch.tensor(Data_Collection_adaptive[:, 0:1], dtype=torch.float32, device=device, requires_grad=True)
            self.G_f = torch.tensor(Data_Collection_adaptive[:, 1:5], dtype=torch.float32, device=device, requires_grad=True)

            # Fine-tuning with Adam optimizer
            print('Adam optimization begins')
            start_time = time.time()
            for it_Adam in range(1000):
                self.optimizer_Adam.step(closure_adam)

                if it_Adam % 10 == 0:
                    elapsed = time.time() - start_time
                    loss = closure_adam().item()
                    loss_u = torch.mean((self.u - self.net_u(self.x, self.t)) ** 2).item()
                    loss_f = self.loss_f_coeff * torch.mean(self.net_f(self.x_f, self.t_f, self.G_f)[0] ** 2).item()
                    loss_lambda = (1e-7 * torch.norm(self.lambda1, p=1)).item()
                    lambda1_value = self.lambda1.detach().cpu().numpy()

                    print('It: %d, Log Loss: %.3e, loss_u: %e, loss_f: %e, loss_lambda: %e, Time: %.2f' % 
                          (it_Adam, loss, loss_u, loss_f, loss_lambda, elapsed))

                    lamu = self.lambda1.detach().cpu().numpy()

                    # Calculate validation losses
                    loss_u_val = torch.mean((self.u_val - self.net_u(self.x_val, self.t_val)) ** 2).item()
                    loss_f_val = self.loss_f_coeff * torch.mean(self.net_f(self.x_val, self.t_val, self.G_val)[0] ** 2).item()
                    loss_val = math.log(loss_u_val + loss_f_val)

                    # Update global loss histories
                    global loss_history_Adam, lambda_history_Adam, loss_u_history_Adam
                    global loss_f_history_Adam, loss_lambda_history_Adam
                    global loss_history_Adam_val, loss_u_history_Adam_val, loss_f_history_Adam_val

                    loss_history_Adam = np.append(loss_history_Adam, loss)
                    lambda_history_Adam = np.append(lambda_history_Adam, lambda1_value, axis=1)
                    loss_u_history_Adam = np.append(loss_u_history_Adam, loss_u)
                    loss_f_history_Adam = np.append(loss_f_history_Adam, loss_f)
                    loss_lambda_history_Adam = np.append(loss_lambda_history_Adam, loss_lambda)

                    loss_history_Adam_val = np.append(loss_history_Adam_val, loss_val)
                    loss_u_history_Adam_val = np.append(loss_u_history_Adam_val, loss_u_val)
                    loss_f_history_Adam_val = np.append(loss_f_history_Adam_val, loss_f_val)

                    lambda_history_Adam = np.append(lambda_history_Adam, lamu, axis=1)

                    start_time = time.time()
        
        # Final sparse identification
        print('Final STRidge optimization')
        self.callTrainSTRidge() 

        # Post-training optimization
        print('Post-training L-BFGS-B optimization begins')
        non_zero_mask = (self.lambda1 != 0).squeeze()
        self.lambda_nonzero = nn.Parameter(self.lambda1.data[non_zero_mask].clone())
        
        # Configure final optimizer
        self.parameters_pt = self.weights + self.biases + [self.lambda1]
        self.optimizer_Post = torch.optim.LBFGS(
            self.parameters_pt,
            max_iter=20000,
            history_size=50,
            max_eval=20000,
            tolerance_grad=1e-7,
            tolerance_change=1e-10,
            line_search_fn="strong_wolfe"
        )
        
        # Define closure function for post-training
        def closure_pt():
            self.optimizer_Post.zero_grad()
            loss, loss_u, loss_f, loss_lambda, loss_val, loss_u_val, loss_f_val = self.loss_function()
            loss.backward()
            
            # Zero out gradients for parameters that were thresholded to zero
            with torch.no_grad():
                zero_mask = ~non_zero_mask
                self.lambda1.grad[zero_mask] = 0
                
            self.callback_Post(loss, loss_u, loss_f, loss_lambda, loss_val, loss_u_val, loss_f_val, self.lambda1)
            return loss
        
        self.optimizer_Post.step(closure_pt)

    def predict(self, X_star):
        """Make predictions using the trained model."""
        x_star = torch.tensor(X_star[:, 0:1], dtype=torch.float32, device=device)
        t_star = torch.tensor(X_star[:, 1:2], dtype=torch.float32, device=device)

        u_star = self.net_u(x_star, t_star).detach().cpu().numpy()
        return u_star

    def callTrainSTRidge(self):
        """Wrapper for the STRidge algorithm to identify sparse coefficients."""
        lam = 1e-5
        d_tol = dtol
        maxit = 100  # Maximum iterations for tolerance search
        STR_iters = 10  # Maximum iterations for STRidge
        
        l0_penalty = None
        normalize = 2
        split = 0.8
        print_best_tol = False
        
        # Get current SINDy components
        f_pred, Phi_pred, u_t_pred = self.net_f(self.x_f, self.t_f, self.G_f)
        
        # Run STRidge algorithm
        lambda2 = self.TrainSTRidge(Phi_pred.detach().cpu().numpy(), u_t_pred.detach().cpu().numpy(), 
                                    lam, d_tol, maxit, STR_iters, l0_penalty, normalize, split, print_best_tol)

        # Update model coefficients
        self.lambda1.data = torch.tensor(lambda2, dtype=torch.float32, device=device)

    def TrainSTRidge(self, R0, Ut, lam, d_tol, maxit, STR_iters=10, l0_penalty=None, normalize=2, split=0.8, print_best_tol=False):
        """
        Train SINDy model using STRidge (Sparse Ridge Regression).
        
        Implements the algorithm from:
        Rudy, Samuel H., et al. "Data-driven discovery of partial differential equations."
        Science Advances 3.4 (2017): e1602614.
        """
        # Normalize data
        n, d = R0.shape
        R = np.zeros((n, d), dtype=np.float32)
        
        if normalize != 0:
            Mreg = np.zeros((d, 1))
            for i in range(d):
                Mreg[i] = 1.0 / (np.linalg.norm(R0[:, i], normalize))
                R[:, i] = Mreg[i] * R0[:, i]
            normalize_inner = 0
        else:
            R = R0
            Mreg = np.ones((d, 1)) * d
            normalize_inner = 2

        # Track normalized coefficients
        global lambda_normalized_history_STRidge
        lambda_normalized_history_STRidge = np.append(lambda_normalized_history_STRidge, Mreg, axis=1)

        # Split data into training and testing sets
        np.random.seed(0)  # For reproducibility
        n, _ = R.shape
        train = np.random.choice(n, int(n * split), replace=False)
        test = [i for i in np.arange(n) if i not in train]
        TrainR = R[train, :]
        TestR = R[test, :]
        TrainY = Ut[train, :]
        TestY = Ut[test, :]
        
        # Initialize tolerance
        d_tol = float(d_tol)
        if self.it == 0:
            self.tol = d_tol
            
        # Initialize with current coefficients
        w_best = self.lambda1.detach().cpu().numpy() / Mreg
        
        # Calculate initial error
        err_f = np.mean((TestY - TestR.dot(w_best)) ** 2)
        
        # Set up L0 penalty
        if l0_penalty is None and self.it == 0:
            self.l0_penalty_0 = err_f
            l0_penalty = self.l0_penalty_0
        elif l0_penalty is None:
            l0_penalty = self.l0_penalty_0
            
        # Composite error with regularization
        err_lambda = l0_penalty * np.count_nonzero(w_best)
        err_best = err_f + err_lambda
        tol_best = 0

        # Update loss histories
        global loss_history_STRidge, loss_f_history_STRidge, loss_lambda_history_STRidge, tol_history_STRidge
        loss_history_STRidge = np.append(loss_history_STRidge, err_best)
        loss_f_history_STRidge = np.append(loss_f_history_STRidge, err_f)
        loss_lambda_history_STRidge = np.append(loss_lambda_history_STRidge, err_lambda)
        tol_history_STRidge = np.append(tol_history_STRidge, tol_best)

        # Adaptively adjust tolerance to find optimal sparsity
        for iter in range(maxit):
            # Get coefficients with current tolerance
            w = self.STRidge(TrainR, TrainY, lam, STR_iters, self.tol, Mreg, normalize=normalize_inner)

            # Calculate errors
            err_f = np.mean((TestY - TestR.dot(w)) ** 2)
            err_lambda = l0_penalty * np.count_nonzero(w)
            err = err_f + err_lambda

            # Update if current model is better
            if err <= err_best:
                err_best = err
                w_best = w
                tol_best = self.tol
                self.tol = self.tol + d_tol

                # Update histories
                loss_history_STRidge = np.append(loss_history_STRidge, err_best)
                loss_f_history_STRidge = np.append(loss_f_history_STRidge, err_f)
                loss_lambda_history_STRidge = np.append(loss_lambda_history_STRidge, err_lambda)
                tol_history_STRidge = np.append(tol_history_STRidge, tol_best)
            else:
                # Reduce tolerance increment
                self.tol = max([0, self.tol - 2 * d_tol])
                d_tol = d_tol / 1.618
                self.tol = self.tol + d_tol

        if print_best_tol:
            print("Optimal tolerance:", tol_best)

        # Track optimal tolerance
        global optimaltol_history
        optimaltol_history = np.append(optimaltol_history, tol_best)

        return np.real(np.multiply(Mreg, w_best))

    def STRidge(self, X0, y, lam, maxit, tol, Mreg, normalize=2, print_results=False):
        """
        Sparse Ridge Regression algorithm for identifying sparse coefficient vectors.
        
        Iteratively applies ridge regression while thresholding small coefficients
        to zero to promote sparsity in the solution.
        """
        n, d = X0.shape
        X = np.zeros((n, d), dtype=np.complex64)
        
        # Normalize data if requested
        if normalize != 0:
            Mreg = np.zeros((d, 1))
            for i in range(d):
                Mreg[i] = 1.0 / (np.linalg.norm(X0[:, i], normalize))
                X[:, i] = Mreg[i] * X0[:, i]
        else:
            X = X0

        # Initialize with current coefficients
        w = self.lambda1.detach().cpu().numpy() / Mreg
        
        # Identify significant coefficients
        biginds = np.where(abs(w) > tol)[0]
        num_relevant = d
        
        # Track ridge regression iterations
        global ridge_append_counter_STRidge
        ridge_append_counter = 0
        
        # Record coefficient history
        global lambda_history_STRidge
        lambda_history_STRidge = np.append(lambda_history_STRidge, np.multiply(Mreg, w), axis=1)
        ridge_append_counter += 1

        # Iterative thresholding
        for j in range(maxit):
            # Identify small coefficients to threshold
            smallinds = np.where(abs(w) < tol)[0]
            new_biginds = [i for i in range(d) if i not in smallinds]

            # Check if convergence reached
            if num_relevant == len(new_biginds):
                break
            else:
                num_relevant = len(new_biginds)
            
            # Handle case where all coefficients are thresholded
            if len(new_biginds) == 0:
                if j == 0:
                    if normalize != 0:
                        lambda_history_STRidge = np.append(lambda_history_STRidge, w * Mreg, axis=1)
                        ridge_append_counter += 1
                        ridge_append_counter_STRidge = np.append(ridge_append_counter_STRidge, ridge_append_counter)
                        return np.multiply(Mreg, w)
                    else:
                        lambda_history_STRidge = np.append(lambda_history_STRidge, w * Mreg, axis=1)
                        ridge_append_counter += 1
                        ridge_append_counter_STRidge = np.append(ridge_append_counter_STRidge, ridge_append_counter)
                        return w
                else:
                    break
            
            # Update significant coefficient indices
            biginds = new_biginds
            
            # Threshold small coefficients to zero
            w[smallinds] = 0
            
            # Solve ridge regression for remaining coefficients
            if lam != 0:
                w[biginds] = np.linalg.lstsq(
                    X[:, biginds].T.dot(X[:, biginds]) + lam * np.eye(len(biginds)), 
                    X[:, biginds].T.dot(y)
                )[0]
                lambda_history_STRidge = np.append(lambda_history_STRidge, np.multiply(Mreg, w), axis=1)
                ridge_append_counter += 1
            else:
                w[biginds] = np.linalg.lstsq(X[:, biginds], y)[0]
                lambda_history_STRidge = np.append(lambda_history_STRidge, np.multiply(Mreg, w), axis=1)
                ridge_append_counter += 1
        
        # Final least squares fit with identified sparsity pattern
        if biginds != []:
            w[biginds] = np.linalg.lstsq(X[:, biginds], y)[0]
        
        # Apply normalization and return
        if normalize != 0:
            lambda_history_STRidge = np.append(lambda_history_STRidge, w * Mreg, axis=1)
            ridge_append_counter += 1
            ridge_append_counter_STRidge = np.append(ridge_append_counter_STRidge, ridge_append_counter)
            return np.multiply(Mreg, w)
        else:
            lambda_history_STRidge = np.append(lambda_history_STRidge, w * Mreg, axis=1)
            ridge_append_counter += 1
            ridge_append_counter_STRidge = np.append(ridge_append_counter_STRidge, ridge_append_counter)
            return w


#%% Main execution
start_time = time.time()

# Hyperparameter configuration
layers = [2, 20, 20, 20, 20, 20, 20, 1]
Pre_step_list = [20000]
dtol_list = [1]
collection_multiple_list = [50]
h_alpha_list = [0.05]

# Project configuration
project_root = os.path.abspath(os.path.join(os.getcwd(), os.pardir)) 
directory_path = f'E:/Batch_OscillatoryShear_7/(2)_run_20250326/18_0328/'
today_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

# Initialize data handlers
GD = GenerateData()
reader = ReadFolder()

# Get dataset classifications
classification_results = reader.return_dimensionless_cases_7items(directory_path)

# Hyperparameter grid search
for dtol in dtol_list:
    for Pre_step in Pre_step_list:
        for collection_multiple in collection_multiple_list:
            for h_alpha in h_alpha_list:
                # Create result directory
                store_excel_path = f"{directory_path}/Results_{today_datetime}（Residual-based Adaptive Sampling）CM={collection_multiple}__α={h_alpha}_dtol={dtol}（Pretrain_{Pre_step}）"
                
                # Process each dataset category
                for item in classification_results:
                    # Clean up previous model if exists
                    try:
                        del model
                        print("Previous model deleted.")
                    except NameError:
                        print("No existing model to delete.")
                    
                    # Reset loss history arrays
                    loss_history_Adam = np.array([0])
                    loss_u_history_Adam = np.array([0])
                    loss_f_history_Adam = np.array([0])
                    loss_lambda_history_Adam = np.array([0])
                    lambda_history_Adam = np.zeros((num_coe, 1))
                    loss_history_Adam_val = np.array([0])
                    loss_u_history_Adam_val = np.array([0])
                    loss_f_history_Adam_val = np.array([0])
                
                    loss_history_STRidge = np.array([0])
                    loss_f_history_STRidge = np.array([0])
                    loss_lambda_history_STRidge = np.array([0])
                    optimaltol_history = np.array([0])
                    tol_history_STRidge = np.array([0])
                    lambda_normalized_history_STRidge = np.zeros((num_coe, 1))
                
                    lambda_history_STRidge = np.zeros((num_coe, 1))
                    ridge_append_counter_STRidge = np.array([0])
                
                    loss_history_Pretrain = np.array([0])
                    loss_u_history_Pretrain = np.array([0])
                    loss_f_history_Pretrain = np.array([0])
                    loss_lambda_history_Pretrain = np.array([0])
                    loss_history_val_Pretrain = np.array([0])
                    loss_u_history_val_Pretrain = np.array([0])
                    loss_f_history_val_Pretrain = np.array([0])
                    step_Pretrain = 0
                
                    lambda_history_Pretrain = np.zeros((num_coe, 1))
                    
                    # Post-training loss histories
                    loss_history_Pt = np.array([0])
                    loss_u_history_Pt = np.array([0])
                    loss_f_history_Pt = np.array([0])
                    loss_lambda_history_Pt = np.array([0])
                    loss_history_val_Pt = np.array([0])
                    loss_u_history_val_Pt = np.array([0])
                    loss_f_history_val_Pt = np.array([0])
                    step_Pt = 0
                    
                    lambda_history_Pt = np.zeros((num_coe, 1))
                    
                    # Reset random seeds for reproducibility
                    np.random.seed(1234)
                    torch.manual_seed(1234)
                    torch.cuda.manual_seed_all(42)
                    
                    # Create storage directory
                    store_path = f"{directory_path}/Results_{today_datetime}（Residual-based Adaptive Sampling）CM={collection_multiple}__α={h_alpha}_dtol={dtol}（Pretrain_{Pre_step}）/{item['category']}"
                    
                    if not os.path.exists(store_path):
                        os.makedirs(store_path)
                        print(f"Created directory: '{store_path}'")
                    else:
                        print(f"Directory exists: '{store_path}'")
                        
                    # Get dataset names for current category
                    name_list = item['folders']
                    
                    # Generate and prepare data
                    Data_Total, Data_Train, Data_Validation, Data_Collection, lb, ub, t = \
                        GD.create_DimensionalCase_data(name_list, project_root, collection_multiple, directory_path, log_flag=False)
                    
                    # Initialize and train model
                    model = PhysicsInformedNN(Data_Train, Data_Collection, Data_Validation, layers, lb, ub,
                                              name_list, project_root, collection_multiple, directory_path, h_alpha, Pre_step, dtol, log_flag=False)
                    model.train(10)  # Run 10 ADO iterations
                    
                    # Validate and visualize results
                    X_u_meas = np.vstack((Data_Train[:, 5], Data_Train[:, 0])).T
                    u_meas = Data_Train[:, 4].reshape(-1, 1)
                    
                    # Extract variables for analysis
                    t = Data_Total[:, 0]
                    γ = Data_Total[:, 1]
                    γ_t = Data_Total[:, 2]
                    γ_tt = Data_Total[:, 3]
                    σ = Data_Total[:, 4]
                    abs_γ_t = abs(γ_t)
                    abs_γ_tt = abs(γ_tt)
                    
                    # Plot true vs predicted values
                    xx = torch.tensor(Data_Validation[:, 5:6], dtype=torch.float32, device=device, requires_grad=True)
                    tt = torch.tensor(Data_Validation[:, 0:1], dtype=torch.float32, device=device, requires_grad=True)
                    u_validation_pred = model.net_u(xx, tt)
                    u_validation_true = Data_Validation[:, 4:5]
                    
                    # Configure plot styling
                    mpl.rcParams['font.family'] = 'Times New Roman'
                    mpl.rcParams['font.serif'] = 'Times New Roman'
                    plt.figure(figsize=(10, 6)) 
                    plt.scatter(tt.detach().cpu().numpy(), u_validation_true, label='σ_DEM', color='blue', s=10, alpha=0.99)
                    plt.scatter(tt.detach().cpu().numpy(), u_validation_pred.detach().cpu().numpy(), label='σ_NN', color='red', marker='x', s=10, alpha=0.6)
                    plt.xlabel('t (s)', fontsize=18)
                    plt.ylabel(r'$\sigma_{xy}$/$P_0$', fontsize=18)
                    plt.tick_params(axis='both', which='major', labelsize=12)
                    plt.legend(fontsize=14)
                    plt.savefig(f'{store_path}/NN_Predicted_Result.png')
                    
                    # Analyze and plot time derivatives
                    xx0 = torch.tensor(Data_Total[:, 5:6], dtype=torch.float32, device=device, requires_grad=True)
                    tt0 = torch.tensor(Data_Total[:, 0:1], dtype=torch.float32, device=device, requires_grad=True)
                    u_total_pred0 = model.net_u(xx0, tt0)
                    u_total_grad0 = torch.autograd.grad(u_total_pred0, tt0, grad_outputs=torch.ones_like(u_total_pred0), create_graph=True)[0] 
                    
                    plt.figure()
                    plt.scatter(tt0.detach().cpu().numpy(), u_total_grad0.detach().cpu().numpy(), label='σ_t (AD)', color='blue', s=10)
                    plt.xlabel('t (s)')
                    plt.ylabel(r'$\partial \sigma_{xy} / \partial t$')
                    plt.legend()
                    
                    # Plot SINDy predicted vs actual derivatives
                    xx1 = torch.tensor(Data_Total[:, 5:6], dtype=torch.float32, device=device, requires_grad=True)
                    tt1 = torch.tensor(Data_Total[:, 0:1], dtype=torch.float32, device=device, requires_grad=True)
                    u_total_pred = model.net_u(xx1, tt1)
                    u_total_grad = torch.autograd.grad(u_total_pred, tt1, grad_outputs=torch.ones_like(u_total_pred), create_graph=True)[0] 
                    
                    # Calculate SINDy prediction
                    σ_t = u_total_grad.squeeze(dim=1).detach().cpu().numpy()
                    ones_tensor = np.ones((t.shape[0], ))
                    
                    # Reconstruct feature library
                    features = [ones_tensor, σ, γ_t, γ_tt, σ*σ, γ_t*γ_t, γ_tt*γ_tt, abs_γ_t, abs_γ_tt,
                            σ*γ_t, σ*γ_tt, σ*abs_γ_t, σ*abs_γ_tt, abs_γ_t*γ_t, abs_γ_t*γ_tt,
                            (σ*σ)*γ_t, (σ*σ)*γ_tt, (σ*σ)*abs_γ_t, (σ*σ)*abs_γ_tt, γ_t*(γ_tt*γ_tt), 
                            γ_tt*(γ_t*γ_t), σ_t*γ_t, σ_t*abs_γ_t]
                    
                    # Get final coefficients
                    lambda1_value = model.lambda1.detach().cpu().numpy()
                    lambda1_true = np.zeros((num_coe, 1))
                    lambda1_true = np.ravel(lambda1_true)
                    lambda1_value = np.ravel(lambda1_value)
                    
                    # Make SINDy predictions
                    Phi = np.vstack(features).T
                    σ_t_pred = np.dot(Phi, lambda1_value)
                    
                    # Plot comparison
                    plt.figure(figsize=(10, 6))
                    plt.scatter(tt1.detach().cpu().numpy(), u_total_grad.detach().cpu().numpy(), label=r'$\mu_t$_AutoDiff', color='blue', s=10)
                    plt.scatter(tt1.detach().cpu().numpy(), σ_t_pred, label=r'$\mu_t$_Equation', color='red', marker='x', s=10, alpha=0.6)
                    plt.xlabel('t (s)')
                    plt.ylabel(r'$\mu_t$')
                    plt.legend()
                    plt.savefig(f'{store_path}\\σ_t_Predicted_Result.png')
                    
                    # Save results and diagnostics
                    store_path = f'{store_path}/'
                    
                    # Write error metrics to file
                    f = open(f'{store_path}\\stdout.txt', "a+")
                    
                    X_u_train = torch.tensor(np.concatenate((Data_Train[:, 1:2], Data_Train[:, 0:1]), axis=1))
                    u_train = torch.tensor(np.concatenate((Data_Train[:, 4:5], Data_Train[:, 0:1]), axis=1))
                    
                    X_u_val = torch.tensor(np.concatenate((Data_Validation[:, 1:2], Data_Validation[:, 0:1]), axis=1))
                    u_val = torch.tensor(np.concatenate((Data_Validation[:, 4:5], Data_Validation[:, 0:1]), axis=1))
                    
                    u_train_Pred = model.predict(X_u_train)
                    Error_u_Train = np.linalg.norm(u_train - u_train_Pred, 2) / np.linalg.norm(u_train, 2)
                    f.write('Training Error u: %e \n' % (Error_u_Train))
                    
                    u_val_Pred = model.predict(X_u_val)
                    Error_u_Val = np.linalg.norm(u_val - u_val_Pred, 2) / np.linalg.norm(u_val, 2)
                    f.write('Validation Error u: %e \n' % (Error_u_Val))
                    
                    elapsed = time.time() - start_time
                    f.write('Training time: %.4f \n' % (elapsed))
                    f.close()
                    
                    with open(f'{store_path}\\output.txt', 'w') as f:
                        # Construct discovered equation string
                        disc_eq_temp = []
                        for i_lib in range(len(model.library_description)):
                            if lambda1_value[i_lib] != 0:
                                disc_eq_temp.append(f"{lambda1_value[i_lib]:.4f}{model.library_description[i_lib]}")
                        disc_eq = '+'.join(disc_eq_temp)
                        f.write('The discovered equation: σ_t = ' + disc_eq)
                    
                    # Plot final coefficients
                    fig = plt.figure()
                    plt.plot(lambda1_value, label='Identified coefficients')
                    plt.title('SINDy Coefficients')
                    plt.legend()
                    plt.savefig(f'{store_path}\\identified_coeff.png')
                
                    # Save coefficients to Excel
                    excel_path = f'{store_excel_path}\\SINDy_Coefficients.xlsx'
                    library_descriptions = model.library_description
                    lambda_values = lambda1_value
                    
                    if os.path.exists(excel_path):
                        wb = load_workbook(excel_path)
                        ws = wb.active
                    else:
                        df = pd.DataFrame([library_descriptions], index=['Library Description'])
                        df.to_excel(excel_path, index=True, engine='openpyxl')
                        wb = load_workbook(excel_path)
                        ws = wb.active
                    
                    current_row = ws.max_row
                    lambda_values_list = lambda_values.tolist()
                    index_name = f"{item['category']}"
                    lambda_values_list.insert(0, index_name)
                    ws.append(lambda_values_list)
                    
                    # Format Excel cells
                    number_format = NamedStyle(name='number_format', number_format='0.000')
                    alignment = Alignment(horizontal='center', vertical='center')
                    for row in ws.iter_rows(min_row=current_row + 1, max_row=current_row + 1):
                        for cell in row:
                            cell.alignment = alignment
                    
                    # Adjust column widths
                    def adjust_column_width(ws):
                        for col in ws.columns:
                            max_length = 0
                            column = col[0].column_letter
                            for cell in col:
                                try:
                                    if cell.value and len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = max_length + 2
                            ws.column_dimensions[column].width = adjusted_width
                    
                    adjust_column_width(ws)
                    wb.save(excel_path)
                    
                    # Clean up
                    del model
                    
                    # Save comprehensive results for analysis
                    end_time_total = time.time()
                    run_time = end_time_total - start_time_total
                    
                    # Collect variables for saving
                    variables = {
                        'lambda_history_Adam': lambda_history_Adam,
                        'lambda_history_Pretrain': lambda_history_Pretrain,
                        'lambda_history_Pt': lambda_history_Pt,
                        'lambda_history_STRidge': lambda_history_STRidge,
                        'lambda_normalized_history_STRidge': lambda_normalized_history_STRidge,
                        'lambda_values': lambda_values,
                        'optimaltol_history': optimaltol_history,
                        'tol_history_STRidge': tol_history_STRidge,
                        'run_time': run_time
                    }
                    
                    # Add all loss variables
                    local_vars = list(locals().keys())
                    for var_name in local_vars:
                        if var_name.startswith('loss'):
                            variables[var_name] = locals()[var_name]
                    
                    # Save to .mat file for further analysis
                    scipy.io.savemat(f'{store_path}\\ErrorAnalysis.mat', variables)

